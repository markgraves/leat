"""Configuration data for search"""

from collections import defaultdict
import csv
import json
from pathlib import Path
import re
from typing import NamedTuple, Optional, Union, Sequence

from .predefined_configurations import PredefinedConfigurations

try:
    import openpyxl

    EXCEL_READ_AVAILABLE = True
    """True iff Excel reading is supported"""
except ModuleNotFoundError:
    print(
        "WARNING:",
        "Excel config file reading not available. Need to: pip install openpyxl",
    )
    EXCEL_READ_AVAILABLE = False


class ConfigData:
    """Configuration data for search

    Attributes:
      config_file: str | Path | None: Path to the configuration file
      short_name: str: Short name of the configuration, e.g, name of predefined configuration or stem of filename
      data: dict: Configuration data loaded from the file

    Example:
      `ConfigData(config_file="SearchConfig.xlsx")`
      `ConfigData(predefined_configuration='BasicSearch')`

    """

    def __init__(
        self,
        config_file: Optional[Union[str, Path]] = None,
        save_json_if_outdated: bool = True,
        load_json_if_current: bool = True,
        json_config_file: Optional[Union[str, Path]] = None,
        predefined_configuration: Optional[str] = None,
    ):
        """
        Creates ConfigData by loading a user-specified or predefined configuration file.
        User files can be Excel, CSV, or a json version of ConfigData

        Args:
          config_file: str | Path | None: Path to the configuration file  (Default value = None)
          save_json_if_outdated: bool: Save fresh json if the config file is new (Default value = True)
          load_json_if_current: bool: Load json file if as new as config file (Default value = True)
          json_config_file: str | Path | None: Path to the json file (Default value = None)
          predefined_configuration: str | None: Name of a predefined configuration (Default value = None)

        Note: Predefined configurations are in core.predefined_configurations

        Example:

        """
        self.data: dict = {}
        self.config_file = None
        self.short_name: str = "Empty"
        if predefined_configuration:
            config_file = PredefinedConfigurations.data.get(predefined_configuration)
            if config_file:
                if config_file.exists():
                    self.load_config_file(config_file, json_config_file=None)
                    self.short_name = "#" + predefined_configuration
                else:
                    print(
                        "ERROR:",
                        "Internal error. Known configuration named",
                        predefined_configuration,
                        "not found at",
                        config_file,
                    )
            else:
                print("ERROR:", "Unknown configuration name:", predefined_configuration)
        elif config_file is None and json_config_file is None:
            pass
        else:
            self.load_config_file(
                config_file,
                save_json_if_outdated,
                load_json_if_current,
                json_config_file,
            )

    def load_config_file(
        self,
        config_file: Optional[Union[str, Path]] = None,
        save_json_if_outdated: bool = True,
        load_json_if_current: bool = True,
        json_config_file: Optional[Union[str, Path]] = None,
    ):
        """
        Loads ConfigData from either an Excel or CSV configuration file, or a saved json version of that data

        Args:
          config_file: str | Path | None: Path to the configuration file  (Default value = None)
          save_json_if_outdated: bool: Save fresh json if the config file is new (Default value = True)
          load_json_if_current: bool: Load json file if as new as config file (Default value = True)
          json_config_file: str | Path | None: Path to the json file (Default value = None)
        """
        if not config_file:
            if json_config_file:
                return self.load_config_file_json(json_config_file)
            print("ERROR:", "No config file specified to load.")
            return
        config_file = Path(config_file)
        config_file_type = self.get_config_file_type(config_file)
        if config_file_type == "csv":
            return self.load_config_file_csv(config_file)
        if config_file_type == "json":
            return self.load_config_file_json(config_file)
        if config_file_type != "xlsx":
            print(
                "ERROR:",
                "Unknown config file type",
                config_file_type,
                "for load file:",
                config_file,
            )
            return
        if json_config_file:
            json_config_file = Path(json_config_file)
        else:
            json_config_file = json_config_file_for_excel(config_file)
        if (
            json_config_file
            and json_config_file.exists()
            and config_file.exists()
            and json_config_file.stat().st_mtime > config_file.stat().st_mtime
        ):
            if load_json_if_current:
                return self.load_config_file_json(json_config_file)
            self.load_config_file_xlsx(config_file)
            return
        if self.load_config_file_xlsx(config_file) and save_json_if_outdated:
            self.write_config_data_json(json_config_file)

    @staticmethod
    def get_config_file_type(filename: [Union[str, Path]]):
        """
        Returns the type (extension) of a configuration file, i.e., xlsx, csv, json

        Args:
          filename: str | Path: Path to the configuration file

        Returns:
          str: Type of the file (extension)
        """
        fileext = Path(filename).suffix.strip(".")
        return fileext

    def load_config_file_csv(self, filename: [Union[str, Path]]) -> bool:
        """
        Read csv configuration file, storing the resulting config dict in the instance

        Args:
          filename: str | Path: Path to the configuration file

        Returns:
          bool: True if file was loaded, otherwise `open` will error
        """
        print("INFO:", "Loading csv config file:", filename)
        with open(filename, "r", encoding="utf-8-sig") as ifp:
            reader = csv.DictReader(ifp)
            result = defaultdict(list)
            for row in reader:
                for k, v in row.items():
                    if v:
                        result[k].append(v.strip())
        self.data = dict(result)
        self.config_file = filename
        self.short_name = filename.stem
        return True

    def load_config_file_xlsx(self, filename: [Union[str, Path]]) -> bool:
        """
        Load Excel xlsx configuration file

        Args:
          filename: str | Path: Path to the configuration file

        Returns:
          bool: True if file was loaded, otherwise `openpyxl` will error
        """
        print("INFO:", "Loading xlsx config file:", filename)
        wb = openpyxl.load_workbook(filename, data_only=True)
        result = {}
        for sn in wb.sheetnames:
            sheet_type = self.get_sheetname_type(sn)
            if sheet_type == "SEARCH":
                result[sn.strip()] = read_config_sheet_search(wb[sn])
            elif sheet_type == "PATTERN":
                result[sn.strip()] = read_config_sheet_pattern(wb[sn])
        self.data = dict(result)
        self.config_file = filename
        self.short_name = filename.stem
        return True

    @staticmethod
    def get_sheetname_type(sheet_name: str) -> str:
        """
        Return the type of sheet name, e.g., SEARCH, PATTERN, or UNKNOWN

        Args:
          sheet_name: str: Name of an openpyxl worksheet

        Returns:
          str: Type of the sheet, namely, SEARCH, PATTERN, or UNKNOWN
        """
        lower = sheet_name.casefold()
        if "search" in lower:
            return "SEARCH"
        if "pattern" in lower:
            return "PATTERN"
        return "UNKNOWN"

    def write_config_data_json(self, filename: [Union[str, Path]]):
        """
        Write config information as json

        Args:
          filename: str | Path: Path to the json configuration file

        Returns:
          bool: True if file was created, otherwise `open` will error
        """
        newconfig = {k: config_json_format_simplify(v) for k, v in self.data.items()}
        print("INFO:", "Saving json config file", filename)
        with open(filename, "w") as ofp:
            json.dump(newconfig, ofp, indent=4)

    def load_config_file_json(self, filename: [Union[str, Path]]) -> bool:
        """
        Read config information from json file

        Args:
          filename: str | Path: Path to the json configuration file

        Returns:
          bool: True if file was loaded, otherwise `open` will error
        """
        print("INFO:", "Loading json config file:", filename)
        with open(filename, "r") as ifp:
            data = json.load(ifp)
        # Hooks to correct serialized type changes
        config = {k: config_json_format_restore(v) for k, v in data.items()}
        self.data = config
        self.config_file = filename
        self.short_name = filename.stem
        return True

    def __str__(self):
        """Format instance as a string"""
        return f"<{__class__.__name__} {self.short_name}>"


class PatternConcept(NamedTuple):
    """
    Concept for pattern, which also tracks flag for matching

    Attributes:
      concept: str: Name of the concept
      flags: int: re.FLAGS
    """

    concept: str
    flags: int = 0  # re.NOFLAG


CLEAN_COLUMN_NAMES_MAPPING = {
    "concept": "CONCEPT",
    "entity": "CONCEPT",
    "pattern": "PATTERN",
    "regex": "PATTERN",
    "regular expression": "PATTERN",
    "case insensitive": "CASE INSENSITIVE",
    "case insens": "CASE INSENSITIVE",
    "ignore case": "CASE INSENSITIVE",
}
"""Maps from terms used in column names to clean strings"""


def clean_column_name_index(colnames: Sequence[str]) -> dict:
    """
    Clean and index column names for config sheets

    Args:
      colnames: Sequence[str]:

    Returns:
      dict: Mapping from clean column names to index in the config sheet for that column
    """
    result = {}
    for i, cn in enumerate(colnames):
        cn_lower = cn.casefold().strip()
        for k, v in CLEAN_COLUMN_NAMES_MAPPING.items():
            if k in cn_lower:
                if v not in result:
                    # Take first match, ignoring extra columns in sheet
                    result[v] = i
                break
    return result


def read_config_sheet_search(sheet) -> dict:
    """
    Read config search sheet, returning dict of values

    Args:
      sheet: openpyxl..Worksheet: Worksheet in openpyxl with search terms

    Returns:
      dict: Dictionary mapping column names (concepts) to the (stripped) string values in the column
    """
    sheetvalues = {"_sheet_type": "SEARCH"}
    for col in sheet.iter_cols():
        if col[0].value:
            sheetvalues[col[0].value] = [x.value.strip() for x in col[1:] if x.value]
    return sheetvalues


def read_config_sheet_pattern(sheet) -> dict:
    """
    Read config pattern sheet, returning dict of values

    Args:
      sheet: openpyxl..Worksheet: Worksheet in openpyxl with patterns

    Returns:
      dict: Dictionary mapping concepts (with flags, i.e., PatternConcept) to patterns
    """
    sheetvalues = defaultdict(list)
    sheetvalues["_sheet_type"] = "PATTERN"
    column_name_map = clean_column_name_index(x.value for x in sheet[1])
    for row in sheet.iter_rows(min_row=2):
        if "CONCEPT" not in column_name_map or "PATTERN" not in column_name_map:
            print(
                "Error: Pattern sheet",
                sn,
                "needs concept and pattern columns in",
                str(filename),
            )
        if "CASE INSENSITIVE" in column_name_map:
            val = row[column_name_map["CASE INSENSITIVE"]].value
            if val:
                flags = (
                    re.IGNORECASE if val.casefold().startswith("y") else 0
                )  # re.NOFLAG # noflag in python 3.11
            else:
                # Default is case sensitive for patterns
                flags = 0  # re.NOFLAG # noflag in python 3.11
        else:
            flags = 0  # re.NOFLAG # noflag in python 3.11
        concept_value = row[column_name_map["CONCEPT"]].value
        pattern_value = row[column_name_map["PATTERN"]].value
        if concept_value is None or concept_value == "null" or pattern_value is None:
            if concept_value is None and pattern_value is None:
                # quietly skip blank line
                pass
            elif concept_value is None or concept_value == "null":
                print(
                    "Error: Pattern sheet",
                    sn,
                    "missing concept for pattern=",
                    pattern_value,
                )
            else:
                print(
                    "Error: Pattern sheet",
                    sn,
                    "missing pattern for concept",
                    pattern_value,
                )
            continue
        else:
            sheetvalues[PatternConcept(concept=concept_value, flags=flags)].append(
                pattern_value
            )
    return dict(sheetvalues)


def json_config_file_for_excel(excel_filename: Union[Path, str]):
    """
    Replaces xlsx filename extension with json

    Args:
      excel_filename: str | Path: Path to the Excel configuration file

    Returns:
      path: Path to a similarly located and name file, except with json extension
    """
    return Path(excel_filename).with_suffix(".json")


def config_json_format_simplify(sheet_config: dict) -> dict:
    """
    Hooks to change configuration types for serialization as json

    Changes PatternConcept object to a two-level dict

    Args:
      sheet_config: dict:

    Returns:
      dict: Sheet config with simplified types
    """
    if sheet_config.get("_sheet_type", "") == "PATTERN":
        result = defaultdict(dict)
        for pattern_concept, pats in sheet_config.items():
            if type(pattern_concept) == str and pattern_concept.startswith("_"):
                result[pattern_concept] = pats
                continue
            assert type(pattern_concept) == PatternConcept
            result[pattern_concept.concept][pattern_concept.flags] = pats
        return result
    else:
        return sheet_config


def config_json_format_restore(sheet_config: dict) -> dict:
    """
    Hooks to revert serialized type changes

    Reverses changes in `config_json_format_simplify`

    Args:
      sheet_config: dict: Configuration dict previously prepared for serialization

    Returns:
      dict: With type simplifications undone
    """
    if sheet_config.get("_sheet_type", "") == "PATTERN":
        result = defaultdict(list)
        for concept, flag_sub_dict in sheet_config.items():
            if concept.startswith("_"):
                result[concept] = flag_sub_dict
                continue
            for flags, pats in flag_sub_dict.items():
                result[PatternConcept(concept, int(flags))] = pats
        return dict(result)
    else:
        return sheet_config
